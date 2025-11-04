from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from datetime import datetime, timedelta
from django.utils import timezone
import openpyxl
import csv
import io
import os

from datetime import timedelta, date
from decimal import Decimal
from django.utils import timezone
from django.shortcuts import render
from django.db.models import Sum, Value, F, Q, DecimalField
from django.db.models.functions import Coalesce, Cast, TruncDate

from django.core.files import File
from django.conf import settings
from django.contrib import messages
from datetime import timedelta
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum, F, Value
from django.db.models.functions import Coalesce
from sowafinance.sales.views import _invoice_analytics 
from django.contrib.auth.decorators import login_required
from sowafinance.sales.models import Newinvoice  
from sowafinance.accounts.models import Account,JournalEntry,JournalLine
from sowafinance.sowaf.models import Newcustomer, Newsupplier
from . models import Newcustomer, Newsupplier,Newclient,Newemployee,Newasset

# Create your views here.
# Constants / helpers
INCOME_TYPES  = {"income", "other income"}
EXPENSE_TYPES = {"expense", "other expense", "cost of goods sold"}
BANK_TYPES    = {"bank", "cash and cash equivalents"}
ZERO_DEC      = Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2))
def _to_float(x):
    try:
        return float(x or 0)
    except Exception:
        return 0.0

def home(request):
    today = timezone.localdate()
    start = today - timedelta(days=29)
    end   = today

    # --- Sales series per day (from invoices.total_due) ---
    # Build a per-day bucket in Python (SQLite-safe and simple)
    inv_qs = (
        Newinvoice.objects
        .filter(date_created__gte=start, date_created__lte=end)
        .values("date_created", "total_due")
        .order_by("date_created", "id")
    )
    day_totals = {}
    for row in inv_qs:
        d = row["date_created"]
        day_totals[d] = day_totals.get(d, 0.0) + _to_float(row["total_due"])

    chart_labels = [(start + timedelta(days=i)).strftime("%b %d") for i in range(30)]
    sales_series = [_to_float(day_totals.get(start + timedelta(days=i), 0.0)) for i in range(30)]

    # --- Expenses donut (last 30 days) ---
    jl_exp = (
        JournalLine.objects
        .select_related("entry","account")
        .filter(entry__date__gte=start, entry__date__lte=end,
                account__account_type__in=EXPENSE_TYPES)
        .values("account__account_name")
        .annotate(
            deb=Coalesce(Sum("debit"),  ZERO_DEC),
            cre=Coalesce(Sum("credit"), ZERO_DEC),
        )
        .order_by("account__account_name")
    )
    exp_labels = [r["account__account_name"] for r in jl_exp]
    exp_values = [_to_float((r["deb"] or 0) - (r["cre"] or 0)) for r in jl_exp]

    # --- Bank balances donut (as of end) ---
    bank_labels, bank_values = [], []
    bank_accounts = Account.objects.filter(
        account_type__in=["Bank", "Cash and Cash Equivalents"]
    ).order_by("account_name")
    for acc in bank_accounts:
        agg = acc.journalline_set.filter(entry__date__lte=end).aggregate(
            deb=Coalesce(Sum("debit"), ZERO_DEC),
            cre=Coalesce(Sum("credit"), ZERO_DEC),
        )
        bal = _to_float((agg["deb"] or 0) - (agg["cre"] or 0))
        bank_labels.append(acc.account_name or "Bank")
        bank_values.append(bal)

    # --- Invoices: amounts + counts ---
    paid_amount_dec = (
        Newinvoice.objects
        .filter(date_created__lte=end)
        .aggregate(v=Coalesce(Sum("payments_applied__amount_paid"), ZERO_DEC))
    )["v"] or Decimal("0")
    paid_amount = _to_float(paid_amount_dec)

    total_due_sum = _to_float(Newinvoice.objects.aggregate(v=Coalesce(Sum("total_due"), 0.0))["v"])
    unpaid_amount = max(total_due_sum - paid_amount, 0.0)

    overdue_amount = 0.0
    paid_count = unpaid_count = overdue_count = 0
    for inv in Newinvoice.objects.all().select_related("customer"):
        bal = (inv.total_due or 0.0) - _to_float(inv.amount_paid)  # amount_paid is Decimal via agg
        if bal <= 0.00001:
            paid_count += 1
        elif inv.due_date and inv.due_date < today:
            overdue_amount += _to_float(bal)
            overdue_count  += 1
        else:
            unpaid_count += 1

    inv_amounts = {
        "paid": round(paid_amount, 2),
        "unpaid": round(unpaid_amount, 2),
        "overdue": round(overdue_amount, 2),
    }
    inv_counts = {"paid": paid_count, "unpaid": unpaid_count, "over": overdue_count}

    # --- P&L summary (last 30 days) ---
    jl_30 = JournalLine.objects.select_related("entry","account").filter(
        entry__date__gte=start, entry__date__lte=end
    )
    inc = jl_30.filter(account__account_type__in=INCOME_TYPES).aggregate(
        deb=Coalesce(Sum("debit"), ZERO_DEC), cre=Coalesce(Sum("credit"), ZERO_DEC)
    )
    income = _to_float((inc["cre"] or 0) - (inc["deb"] or 0))

    exp = jl_30.filter(account__account_type__in=EXPENSE_TYPES).aggregate(
        deb=Coalesce(Sum("debit"), ZERO_DEC), cre=Coalesce(Sum("credit"), ZERO_DEC)
    )
    expense = _to_float((exp["deb"] or 0) - (exp["cre"] or 0))

    pl_summary = {"income": income, "expense": expense, "profit": round(income - expense, 2)}

    context = {
        "range_text": f"{start:%b %d} – {end:%b %d}",
        "chart_labels": chart_labels,
        "sales_series": sales_series,
        "exp_labels": exp_labels, "exp_values": exp_values,
        "bank_labels": bank_labels, "bank_values": bank_values,
        "inv_amounts": inv_amounts, "inv_counts": inv_counts,
        "pl_summary": pl_summary,
    }
    return render(request, "Home.html", context)
def assets(request):
    assets = Newasset.objects.all()
      
    return render(request, 'Assets.html', {'assets':assets})
# assets form 
def add_assests(request):
    if request.method=='POST':
            # getting the supplier by id since its a foreign key
        supplier_id = request.POST.get('supplier')
        supplier=None
        if supplier_id:
            try:
                supplier = Newsupplier.objects.get(pk=supplier_id)
            except Newsupplier.DoesNotExist:
                supplier=None
        
        
        
        asset_name = request.POST.get('asset_name')
        asset_tag = request.POST.get('asset_tag')
        asset_category = request.POST.get('asset_category')
        asset_description = request.POST.get('asset_description')
        department = request.POST.get('department')
        custodian = request.POST.get('custodian')
        asset_status = request.POST.get('asset_status')
        purchase_price = request.POST.get('purchase_price')

        funding_source = request.POST.get('funding_source')
        life_span = request.POST.get('life_span')
        depreciation_method = request.POST.get('depreciation_method')
        residual_value = request.POST.get('residual_value')
        accumulated_depreciation = request.POST.get('accumulated_depreciation')
        remaining_value = request.POST.get('remaining_value')
        asset_account = request.POST.get('asset_account')
        capitalization_date = request.POST.get('capitalization_date')
        cost_center = request.POST.get('cost_center')
        asset_condition = request.POST.get('asset_condition')
        maintenance_schedule = request.POST.get('maintenance_schedule')
        insurance_details = request.POST.get('insurance_details')
        notes = request.POST.get('notes')
        asset_attachments =request.FILES.get('asset_attachments')
# handling the date 
        capitalization_date_str = request.POST.get('capitalization_date')
        capitalization_date = None
        if capitalization_date_str:
            try:
                capitalization_date = datetime.strptime(capitalization_date_str, '%d/%m/%Y')
            except ValueError:
                capitalization_date = None  # Or handle error
# purchase date
        purchase_date_str = request.POST.get('purchase_date')
        purchase_date = None
        if purchase_date_str:
            try:
                purchase_date = datetime.strptime(purchase_date_str, '%d/%m/%Y')
            except ValueError:
                purchase_date = None 

    # waranty date
        warranty_str = request.POST.get('warranty')
        warranty = None
        if warranty_str:
            try:
                warranty = datetime.strptime(warranty_str, '%d/%m/%Y')
            except ValueError:
                warranty = None 

    # saving the assets
        asset = Newasset(asset_name=asset_name,asset_tag=asset_tag,asset_category=asset_category,asset_description=asset_description,department=department,custodian=custodian,asset_status=asset_status,purchase_price=purchase_price,purchase_date=purchase_date,supplier=supplier,warranty=warranty,funding_source=funding_source,life_span=life_span,depreciation_method=depreciation_method,residual_value=residual_value,accumulated_depreciation=accumulated_depreciation,remaining_value=remaining_value,asset_account=asset_account,capitalization_date=capitalization_date,cost_center=cost_center,asset_condition=asset_condition,maintenance_schedule=maintenance_schedule,insurance_details=insurance_details,notes=notes,asset_attachments=asset_attachments,)

        asset.save()
        # adding button save actions
        save_action = request.POST.get('save_action')
        if save_action == 'save&new':
            return redirect('add-asset')
        elif save_action == 'save&close':
            return redirect('assets')
    suppliers = Newsupplier.objects.all()
    return render(request, 'assets_form.html', {'suppliers':suppliers})
# editing assets
def edit_asset(request, pk):
    asset = get_object_or_404(Newasset,pk=pk)
    if request.method=='POST':
        asset.asset_name = request.POST.get('asset_name',asset.asset_name)
        asset.asset_tag = request.POST.get('asset_tag',asset.asset_tag)
        asset.asset_category = request.POST.get('asset_category',asset.asset_category)
        asset.asset_description = request.POST.get('asset_description',asset.asset_description)
        asset.department = request.POST.get('department',asset.department)
        asset.custodian = request.POST.get('custodian',asset.custodian)
        asset.asset_status = request.POST.get('asset_status',asset.asset_status)
        asset.purchase_price = request.POST.get('purchase_price',asset.purchase_price)
        asset.purchase_date = request.POST.get('purchase_date',asset.purchase_date)

        asset.funding_source = request.POST.get('funding_source',asset.funding_source)
        asset.life_span = request.POST.get('life_span',asset.life_span) 
        asset.depreciation_method = request.POST.get('depreciation_method',asset.depreciation_method)
        asset.residual_value = request.POST.get('residual_value',asset.residual_value)
        asset.accumulated_depreciation = request.POST.get('accumulated_depreciation',asset.accumulated_depreciation)
        asset.remaining_value = request.POST.get('remaining_value',asset.remaining_value)
        asset.asset_account = request.POST.get('asset_account',asset.asset_account)
        asset.cost_center = request.POST.get('cost_center',asset.cost_center)
        asset.asset_condition = request.POST.get('asset_condition',asset.asset_condition)
        asset.maintenance_schedule = request.POST.get('maintenance_schedule',asset.maintenance_schedule)
        asset.insurance_details = request.POST.get('insurance_details',asset.insurance_details)
        asset.notes = request.POST.get('notes',asset.notes)
        
        # Handle ForeignKey (supplier)
        supplier_id = request.POST.get('supplier')
        if supplier_id:
            try:
                asset.supplier = Newsupplier.objects.get(pk=supplier_id)
            except Newsupplier.DoesNotExist:
                asset.supplier = None
        # handling the date 
        capitalization_date_str = request.POST.get('capitalization_date')
        if capitalization_date_str:
            
            try:
                asset.capitalization_date = datetime.strptime(capitalization_date_str, '%d/%m/%Y')
            except ValueError:
                pass  # Keep the original value or handle error

        purchase_date_str = request.POST.get('purchase_date')
        if purchase_date_str:
            try:
                asset.purchase_date = datetime.strptime(purchase_date_str, '%d/%m/%Y')
            except ValueError:
                pass  # Keep the original value or handle error

        warranty_str = request.POST.get('warranty')
        if warranty_str:
            try:
                asset.warranty = datetime.strptime(warranty_str, '%d/%m/%Y')
            except ValueError:
                pass  
                
# working on the files
        if 'asset_attachments' in request.FILES:
            asset.asset_attachments = request.FILES['asset_attachments']
        asset.save()

        return redirect('sowaf:assets')
    suppliers = Newsupplier.objects.all()
    return render(request, 'assets_form.html', {'asset': asset,'suppliers': suppliers})
# deleting an asset
def delete_asset(request, pk):
    customer = get_object_or_404(Newasset, pk=pk)
    customer.delete()
    return redirect('sowaf:assets')

# importing assets
def download_assets_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets Template"

    headers = [
        'asset_name','asset_tag','asset_category','asset_description','department','custodian','asset_status','purchase_price','purchase_date','supplier','warranty','funding_source','life_span','depreciation_method','residual_value','accumulated_depreciation','remaining_value','asset_account','capitalization_date','cost_center','asset_condition','maintenance_schedule','insurance_details','notes',
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="assets_template.xlsx"'
        return response
# functions to handle the date formats
# Parse capitalization_date (multiple formats)
def parse_capitalization_date_safe(capitalization_date):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(capitalization_date), fmt).date()
        except (ValueError, TypeError):
            continue
    return None

# Parse purchase_date (multiple formats)
def parse_purchase_date_safe(purchase_date):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(purchase_date), fmt).date()
        except (ValueError, TypeError):
            continue
    return None
# Parse warranty (multiple formats)
def parse_warranty_safe(warranty):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(warranty), fmt).date()
        except (ValueError, TypeError):
            continue
    return None
# actual import
def import_assets(request):
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('sowaf:assets')

    excel_file = request.FILES['excel_file']
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith('.csv'):
            decoded_file = excel_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader)  
            for row in reader:
                (
                    asset_name,asset_tag,asset_category,asset_description,department,custodian,asset_status,purchase_price,purchase_date,supplier,warranty,funding_source,life_span,depreciation_method,residual_value,accumulated_depreciation,remaining_value,asset_account,capitalization_date,cost_center,asset_condition,maintenance_schedule,insurance_details,notes,
                ) = row
                capitalization_date = parse_capitalization_date_safe(capitalization_date)
                purchase_date = parse_purchase_date_safe(purchase_date)
                warranty = parse_warranty_safe(warranty)
                
                asset = Newasset.objects.create(
                    
                    asset_name=asset_name,
                    asset_tag=asset_tag,
                    asset_category=asset_category,
                    asset_description=asset_description,
                    department=department,
                    custodian=custodian,
                    asset_status=asset_status,
                    purchase_price=purchase_price,
                    purchase_date=purchase_date,
                    supplier=supplier,
                    warranty=warranty,
                    funding_source=funding_source,
                    life_span=life_span,
                    depreciation_method=depreciation_method,
                    residual_value=residual_value,
                    accumulated_depreciation=accumulated_depreciation,
                    remaining_value=remaining_value,
                    asset_account=asset_account,
                    capitalization_date=capitalization_date,
                    cost_center=cost_center,
                    asset_condition=asset_condition,
                    maintenance_schedule=maintenance_schedule,
                    insurance_details=insurance_details,
                    notes=notes,
                )

        elif file_name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                asset_name,asset_tag,asset_category,asset_description,department,custodian,asset_status,purchase_price,purchase_date,supplier,warranty,funding_source,life_span,depreciation_method,residual_value,accumulated_depreciation,remaining_value,asset_account,capitalization_date,cost_center,asset_condition,maintenance_schedule,insurance_details,notes,
                ) = row

                capitalization_date = parse_capitalization_date_safe(capitalization_date)
                purchase_date = parse_purchase_date_safe(purchase_date)
                warranty = parse_warranty_safe(warranty)

                asset = Newasset.objects.create(
                    asset_name=asset_name,
                    asset_tag=asset_tag,
                    asset_category=asset_category,
                    asset_description=asset_description,
                    department=department,
                    custodian=custodian,
                    asset_status=asset_status,
                    purchase_price=purchase_price,
                    purchase_date=purchase_date,
                    supplier=supplier,
                    warranty=warranty,
                    funding_source=funding_source,
                    life_span=life_span,
                    depreciation_method=depreciation_method,
                    residual_value=residual_value,
                    accumulated_depreciation=accumulated_depreciation,
                    remaining_value=remaining_value,
                    asset_account=asset_account,
                    capitalization_date=capitalization_date,
                    cost_center=cost_center,
                    asset_condition=asset_condition,
                    maintenance_schedule=maintenance_schedule,
                    insurance_details=insurance_details,
                    notes=notes,
                )
        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect('sowaf:assets')

        messages.success(request, "asset data imported successfully.")
        return redirect('sowaf:assets')

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect('sowaf:assets')

# customer view
def customers(request):
    customers = Newcustomer.objects.all()

    return render(request, 'Customers.html', {'customers':customers})

# customer form view

def add_customer(request):
    if request.method == 'POST':
        logo =request.FILES.get('logo')
        if logo:
            if not logo.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            # restricting the photo size
            if logo.size > 1048576:
                messages.error(request, "logo file size must not exceed 800kps.")
                return redirect(request.path)
        customer_name =request.POST.get('name')
        company_name =request.POST.get('company')
        email =request.POST.get('email')
        phone_number =request.POST.get('phonenum')
        mobile_number =request.POST.get('mobilenum')
        website =request.POST.get('website')
        tin_number =request.POST.get('tin')
        opening_balance =request.POST.get('balance')
        registration_date_str = request.POST.get('today')
        registration_date = None
        if registration_date_str:
            try:
               registration_date = datetime.strptime(registration_date_str, '%d/%m/%Y')
            except ValueError:
               registration_date = None 
               
        street_one =request.POST.get('street1')
        street_two =request.POST.get('street2')
        city =request.POST.get('city')
        province =request.POST.get('province')
        postal_code =request.POST.get('postalcode')
        country =request.POST.get('country')
        actions =request.POST.get('actions')
        notes =request.POST.get('notes')
        attachments =request.FILES.get('attachments')
        new_customer = Newcustomer(logo=logo,customer_name=customer_name,company_name=company_name,email=email,phone_number=phone_number,mobile_number=mobile_number,website=website,tin_number=tin_number,opening_balance=opening_balance,registration_date=registration_date,street_one=street_one,street_two=street_two,city=city,province=province,postal_code=postal_code,country=country,actions=actions,notes=notes,attachments=attachments)
        new_customer.save()
        # adding save actions
        save_action = request.POST.get('save_action')
        if save_action == 'save&new':
            return redirect('add-customer')
        elif save_action == 'save&close':
            return redirect('sowaf:customers')
       
    return render(request, 'customers_form.html', {})
# editing the customer table

def edit_customer(request, pk):
    customer = get_object_or_404(Newcustomer, pk=pk)

    if request.method == 'POST':
        customer.customer_name = request.POST.get('name',customer.customer_name)
        customer.company_name = request.POST.get('company',customer.company_name)
        customer.email = request.POST.get('email', customer.email)
        customer.phone_number = request.POST.get('phonenum',customer.phone_number)
        customer.mobile_number = request.POST.get('mobilenum')
        customer.website = request.POST.get('website',customer.website)
        customer.tin_number = request.POST.get('tin',customer.tin_number)
        customer.opening_balance = request.POST.get('balance',customer.opening_balance)
        registration_date_str = request.POST.get('today')
        if registration_date_str:
            try:
                customer.registration_date = datetime.strptime(registration_date_str, '%d/%m/%Y')
            except ValueError:
                pass  # Keep the original value or handle error
        customer.street_one = request.POST.get('street1',customer.street_one)
        customer.street_two = request.POST.get('street2',customer.street_two)
        customer.city = request.POST.get('city',customer.city)
        customer.province = request.POST.get('province',customer.province)
        customer.postal_code = request.POST.get('postalcode',customer.postal_code)
        customer.country = request.POST.get('country',customer.country)
        customer.actions = request.POST.get('actions',customer.actions)
        customer.notes = request.POST.get('notes',customer.notes)

        logo = request.FILES.get('logo')
        if logo:
            if not logo.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            # restricting the photo size
            if logo.size > 1048576:
                messages.error(request, "logo file size must not exceed 800kps.")
                return redirect(request.path)
            customer.logo = logo
        
        if 'attachments' in request.FILES:

            customer.attachments = request.FILES['attachments']

        customer.save()

        return redirect('sowaf:customers')

    return render(request, 'customers_form.html', {'customer': customer})

# Delete view
def delete_customer(request, pk):
    customer = get_object_or_404(Newcustomer, pk=pk)
    customer.delete()
    return redirect('sowaf:customers')
# importing a customer sheet
# template for the download
def download_customers_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Template"

    headers = [
        'name', 'company', 'email', 'phone', 'mobile', 'website', 'tin', 'balance', 'date_str', 'street1', 'street2', 'city', 'province', 'postal_code', 'country', 'actions', 'notes', 'logo'
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customer_template.xlsx"'
        return response
def import_customers(request):
    
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
        file_name = excel_file.name.lower()

        try:
            if file_name.endswith('.csv'):
                decoded_file = excel_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                next(reader)  # Skip header row

                for row in reader:
                   name, company, email, phone, mobile, website, tin, balance, date_str, street1, street2, city, province, postal_code, country, actions, notes, logo = row
                   Newcustomer.objects.create(
                        customer_name=name,
                        company_name=company,
                        email=email,
                        phone_number=phone,
                        mobile_number=mobile,
                        website=website,
                        tin_number=tin,
                        opening_balance=balance,
                        registration_date=date_str,
                        street_one=street1,
                        street_two=street2,
                        city=city,
                        province=province,
                        postal_code=postal_code,
                        country=country,
                        actions=actions,
                        notes=notes,
                    )
                   if logo:
                        image_path = os.path.join(settings.MEDIA_ROOT, 'uploads', logo)
                        if os.path.exists(image_path):
                            with open(image_path, 'rb') as f:
                                Newcustomer.logo.save(logo, File(f), save=False)
                        else:
                            messages.warning(request, f"Image file '{logo}' not found.")
                            Newcustomer.save()
            
            elif file_name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, company, email, phone, mobile, website, tin, balance, date_str, street1, street2, city, province, postal, country, actions, notes, logo = row
                    Newcustomer.objects.create(
                        customer_name=name,
                        company_name=company,
                        email=email,
                        phone_number=phone,
                        mobile_number=mobile,
                        website=website,
                        tin_number=tin,
                        opening_balance=balance,
                        registration_date=date_str,
                        street_one=street1,
                        street_two=street2,
                        city=city,
                        province=province,
                        postal_code=postal,
                        country=country,
                        actions=actions,
                        notes=notes,
                    )
                    if logo:
                        image_path = os.path.join(settings.MEDIA_ROOT, 'uploads', logo)
                        if os.path.exists(image_path):
                            with open(image_path, 'rb') as f:
                                Newcustomer.logo.save(logo, File(f), save=True)
                        else:
                            messages.warning(request, f"Image file '{logo}' not found.")
                            Newcustomer.save()
            else:
                messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
                return redirect('sowaf:customers')
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")
            return redirect('sowaf:customers')
        return redirect('sowaf:customers')   
# clients view

def clients(request):
    clients = Newclient.objects.all()
    return render(request, 'Clients.html', {'clients':clients})

# client form view
def _parse_date(val):
    if not val:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None
def add_client(request):
    if request.method == 'POST':
        # logo validation (PNG, <=1MB) — keep your existing rules
        logo = request.FILES.get('logo')
        if logo:
            if not logo.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1 * 1024 * 1024:
                messages.error(request, "Logo file size must not exceed 1MB.")
                return redirect(request.path)

        # read fields
        company         = request.POST.get('company')
        phone           = request.POST.get('phone')
        company_email   = request.POST.get('company_email')
        address         = request.POST.get('address')
        country         = request.POST.get('country')
        reg_number      = request.POST.get('reg_number')
        start_date      = _parse_date(request.POST.get('start_date'))  # <-- fixed
        contact_name    = request.POST.get('contact_name')
        position        = request.POST.get('position')
        contact         = request.POST.get('contact')
        contact_email   = request.POST.get('contact_email')
        tin             = request.POST.get('tin')
        credit_limit    = request.POST.get('credit_limit')
        payment_terms   = request.POST.get('payment_terms')
        currency        = request.POST.get('currency')
        industry        = request.POST.get('industry')
        status          = request.POST.get('status')
        notes           = request.POST.get('notes')

        client = Newclient(
            logo=logo, company=company, phone=phone, company_email=company_email,
            address=address, country=country, reg_number=reg_number,
            start_date=start_date, contact_name=contact_name, position=position,
            contact=contact, contact_email=contact_email, tin=tin,
            credit_limit=credit_limit, payment_terms=payment_terms,
            currency=currency, industry=industry, status=status, notes=notes
        )
        client.save()

        # save actions
        save_action = request.POST.get('save_action')
        if save_action == 'save':
            return redirect('sowaf:clients')
        if save_action == 'save&new':
            return redirect('sowaf:add-client')
        elif save_action == 'save&close':
            return redirect('sowaf:clients')
        return redirect('sowaf:clients')

    return render(request, 'Clients_form.html', {})  # create flow

# editing the client

def edit_client(request, pk: int):
    client = get_object_or_404(Newclient, pk=pk)

    if request.method == 'POST':
        # optional logo replacement (keep old if none uploaded)
        logo = request.FILES.get('logo')
        if logo:
            if not logo.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1 * 1024 * 1024:
                messages.error(request, "Logo file size must not exceed 1MB.")
                return redirect(request.path)
            client.logo = logo  # replace

        # update fields
        client.company        = request.POST.get('company') or client.company
        client.phone          = request.POST.get('phone') or client.phone
        client.company_email  = request.POST.get('company_email') or client.company_email
        client.address        = request.POST.get('address') or client.address
        client.country        = request.POST.get('country') or client.country
        client.reg_number     = request.POST.get('reg_number') or client.reg_number
        client.start_date     = _parse_date(request.POST.get('start_date')) or client.start_date
        client.contact_name   = request.POST.get('contact_name') or client.contact_name
        client.position       = request.POST.get('position') or client.position
        client.contact        = request.POST.get('contact') or client.contact
        client.contact_email  = request.POST.get('contact_email') or client.contact_email
        client.tin            = request.POST.get('tin') or client.tin
        client.credit_limit   = request.POST.get('credit_limit') or client.credit_limit
        client.payment_terms  = request.POST.get('payment_terms') or client.payment_terms
        client.currency       = request.POST.get('currency') or client.currency
        client.industry       = request.POST.get('industry') or client.industry
        client.status         = request.POST.get('status') or client.status
        client.notes          = request.POST.get('notes') or client.notes

        client.save()

        save_action = request.POST.get('save_action')
        if save_action == 'save':
            return redirect('sowaf:clients')
        if save_action == 'save&new':

            return redirect('sowaf:add-client')
        elif save_action == 'save&close':

            return redirect('sowaf:clients')

        return redirect('sowaf:edit-client', pk=client.id)

    # GET: render same form, pre-filled
    return render(request, 'Clients_form.html', {"client": client})# client delete view
def delete_client(request, pk):
    client = get_object_or_404(Newclient, pk=pk)
    client.delete()
    return redirect('sowaf:clients')

# importing the client
def download_clients_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "clients Template"

    headers = [
        'company', 'phone', 'company_email', 'address', 'country',
        'registration_number', 'start_date', 'contact_name',
        'position', 'contact', 'contact_email', 'tin', 'credit_limit',
        'payment_terms', 'currency', 'industry', 'status',
         'notes', 'logo'
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="clients_template.xlsx"'
        return response
def handle_logo_upload(client, logo):
    if logo:
        image_path = os.path.join(settings.MEDIA_ROOT, 'uploads', logo)
        if os.path.exists(image_path):
            with open(image_path, 'rb') as f:
                client.logo.save(logo, File(f), save=True)
        else:
            messages.warning(None, f"Image file '{logo}' not found.")


def parse_start_date(value):
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def import_clients(request):
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('sowaf:clients')

    excel_file = request.FILES['excel_file']
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith('.csv'):
            decoded_file = excel_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader)  # Skip header row

            for row in reader:
                (
                    company, phone, company_email, address, country,
                    registration_number, start_date, contact_name,
                    position, contact, contact_email, tin, credit_limit,
                    payment_terms, currency, industry, status,
                    notes, logo
                ) = row

                client = Newclient.objects.create(
                    company=company,
                    phone=phone,
                    company_email=company_email,
                    address=address,
                    country=country,
                    reg_number=registration_number,
                    start_date=parse_start_date(start_date),
                    contact_name=contact_name,
                    position=position,
                    contact=contact,
                    contact_email=contact_email,
                    tin=tin,
                    credit_limit=credit_limit,
                    payment_terms=payment_terms,
                    currency=currency,
                    industry=industry,
                    status=status,
                    notes=notes,
                    logo=logo,
                )
                handle_logo_upload(client, logo)

        elif file_name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                    company, phone, company_email, address, country,
                    registration_number, start_date, contact_name,
                    position, contact, contact_email, tin, credit_limit,
                    payment_terms, currency, industry, status,
                    notes, logo
                ) = row

                client = Newclient.objects.create(
                    company=company,
                    phone=phone,
                    company_email=company_email,
                    address=address,
                    country=country,
                    reg_number=registration_number,
                    start_date=parse_start_date(start_date),
                    contact_name=contact_name,
                    position=position,
                    contact=contact,
                    contact_email=contact_email,
                    tin=tin,
                    credit_limit=credit_limit,
                    payment_terms=payment_terms,
                    currency=currency,
                    industry=industry,
                    status=status,
                    notes=notes,
                    logo=logo,
                )
                handle_logo_upload(client, logo)

        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect('sowaf:clients')

        messages.success(request, "Client data imported successfully.")
        return redirect('sowaf:clients')

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect('sowaf:clients')
def employee(request):
    employees = Newemployee.objects.all()

    return render(request, 'Employees.html', {'employees': employees})
# add employee form 
def add_employees(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        gender = request.POST.get('gender')
        dob_str = request.POST.get('dob')
        dob = None
        if dob_str:
            try:
               dob = datetime.strptime(dob_str, '%d/%m/%Y')
            except ValueError:
               dob = None
        nationality = request.POST.get('nationality')
        nin_number = request.POST.get('nin_number')
        tin_number = request.POST.get('tin_number')
        profile_picture = request.FILES.get('profile_picture')
        if profile_picture:
            if not profile_picture.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the profile_picture.")
                return redirect(request.path)
            if profile_picture.size > 1048576:
                messages.error(request, "profile_picture file size must not exceed 1MB.")
                return redirect(request.path)
        phone_number = request.POST.get('phone_number')
        email_address = request.POST.get('email_address')
        residential_address = request.POST.get('residential_address')
        emergency_person = request.POST.get('emergency_person')
        emergency_contact = request.POST.get('emergency_contact')
        relationship = request.POST.get('relationship')
        job_title = request.POST.get('job_title')
        department = request.POST.get('department')
        employment_type = request.POST.get('employment_type')
        status = request.POST.get('status')
        hire_date_str = request.POST.get('hire_date')
        hire_date = None
        if hire_date_str:
            try:
               hire_date = datetime.strptime(hire_date_str, '%d/%m/%Y')
            except ValueError:
               hire_date = None
        supervisor = request.POST.get('supervisor')
        salary = request.POST.get('salary')
        payment_frequency = request.POST.get('payment_frequency')
        payment_method = request.POST.get('payment_method')
        bank_name = request.POST.get('bank_name')
        bank_account = request.POST.get('bank_account')
        bank_branch = request.POST.get('bank_branch')
        nssf_number = request.POST.get('nssf_number')
        insurance_provider = request.POST.get('insurance_provider')
        taxable_allowances = request.POST.get('taxable_allowances')
        intaxable_allowances = request.POST.get('intaxable_allowances')
        additional_notes = request.POST.get('additional_notes')
        doc_attachments = request.FILES.get('doc_attachments')
        # saving the new employee
        employee = Newemployee(first_name=first_name,last_name=last_name,gender=gender,dob=dob,nationality=nationality,nin_number=nin_number,tin_number=tin_number,profile_picture=profile_picture,phone_number=phone_number,email_address=email_address,residential_address=residential_address,emergency_person=emergency_person,emergency_contact=emergency_contact,relationship=relationship,job_title=job_title,department=department,employment_type=employment_type,status=status,hire_date=hire_date,supervisor=supervisor,salary=salary,payment_frequency=payment_frequency,payment_method=payment_method,bank_name=bank_name,bank_account=bank_account,bank_branch=bank_branch,nssf_number=nssf_number,insurance_provider=insurance_provider,taxable_allowances=taxable_allowances,intaxable_allowances=intaxable_allowances,additional_notes=additional_notes,doc_attachments=doc_attachments,)

        employee.save()
        # adding button save actions
        save_action = request.POST.get('save_action')
        if save_action == 'save&new':
            return redirect('add-employee')
        elif save_action == 'save&close':
            return redirect('sowaf:employees')
    
    return render(request, 'employees_form.html', {})

# editing the employee
def edit_employee(request, pk):
    employee = get_object_or_404(Newemployee, pk=pk)

    if request.method == 'POST':
        employee.first_name = request.POST.get('first_name', employee.first_name)
        employee.last_name = request.POST.get('last_name', employee.last_name)
        employee.gender = request.POST.get('gender', employee.gender)
        dob_str = request.POST.get('dob')
        if dob_str:
            try:
                employee.dob = datetime.strptime(dob_str, '%d/%m/%Y')
            except ValueError:
                pass  # Keep the original value or handle error
        employee.nationality = request.POST.get('nationality', employee.nationality)
        employee.nin_number = request.POST.get('nin_number', employee.nin_number)
        employee.tin_number = request.POST.get('tin_number', employee.tin_number)
        employee.phone_number = request.POST.get('phone_number', employee.phone_number)
        employee.email_address = request.POST.get('email_address', employee.email_address)
        employee.residential_address = request.POST.get('residential_address', employee.residential_address)
        employee.emergency_person = request.POST.get('emergency_person', employee.emergency_person)
        employee.emergency_contact = request.POST.get('emergency_contact', employee.emergency_contact)
        employee.relationship = request.POST.get('relationship', employee.relationship)
        employee.job_title = request.POST.get('job_title', employee.job_title)
        employee.department = request.POST.get('department', employee.department)
        employee.employment_type = request.POST.get('employment_type', employee.employment_type)
        employee.status = request.POST.get('status', employee.status)
        hire_date_str = request.POST.get('hire_date')
        if hire_date_str:
            try:
                employee.hire_date = datetime.strptime(hire_date_str, '%d/%m/%Y')
            except ValueError:
                pass  # Keep the original value or handle error
        employee.department = request.POST.get('department', employee.department)
        employee.supervisor = request.POST.get('supervisor', employee.supervisor)
        employee.salary = request.POST.get('salary', employee.salary)
        employee.payment_frequency = request.POST.get('payment_frequency', employee.payment_frequency)
        employee.payment_method = request.POST.get('payment_method', employee.payment_method)
        employee.bank_name = request.POST.get('bank_name', employee.bank_name)
        employee.bank_account = request.POST.get('bank_account', employee.bank_account)
        employee.bank_branch = request.POST.get('bank_branch', employee.bank_branch)
        employee.nssf_number = request.POST.get('nssf_number', employee.nssf_number)
        employee.insurance_provider = request.POST.get('insurance_provider', employee.insurance_provider)
        employee.taxable_allowances = request.POST.get('taxable_allowances', employee.taxable_allowances)
        employee.intaxable_allowances = request.POST.get('intaxable_allowances', employee.intaxable_allowances)
        employee.additional_notes = request.POST.get('additional_notes', employee.additional_notes)

        # ✅ Only update profile_picture if a new one is uploaded
        profile_picture = request.FILES.get('profile_picture')
        if profile_picture:
            if not profile_picture.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the profile picture.")
                return redirect(request.path)
            if profile_picture.size > 1048576:
                messages.error(request, "Profile picture file size must not exceed 1MB.")
                return redirect(request.path)
            employee.profile_picture = profile_picture
        
        if 'doc_attachments' in request.FILES:
            employee.doc_attachments = request.FILES['doc_attachments']

        employee.save()
        return redirect('sowaf:employees')  # Or wherever your list view is

    return render(request, 'employees_form.html', {'employee': employee})

# employee delete view
def delete_employee(request, pk):
    employee  = get_object_or_404(Newemployee, pk=pk)
    employee.delete()
    return redirect('sowaf:employees')

# importing employees
def download_employees_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees Template"

    headers = [
        'first_name', 'last_name', 'gender', 'dob', 'nationality',
        'nin_number', 'tin_number', 'profile_picture', 'phone_number', 'email_address',
        'residential_address', 'emergency_person', 'emergency_contact', 'relationship',
        'job_title', 'department', 'employment_type', 'status', 'hire_date', 'supervisor',
        'salary', 'payment_frequency', 'payment_method', 'bank_name', 'bank_account',
        'bank_branch', 'nssf_number', 'insurance_provider', 'taxable_allowances',
        'intaxable_allowances', 'additional_notes'
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employees_template.xlsx"'
        return response
def handle_profile_picture_upload(employee, profile_picture):
    if profile_picture:
        image_path = os.path.join(settings.MEDIA_ROOT, 'uploads', profile_picture)
        if os.path.exists(image_path):
            with open(image_path, 'rb') as f:
                employee.profile_picture.save(profile_picture, File(f), save=True)
        else:
            messages.warning(None, f"Image file '{profile_picture}' not found.")


# Parse DOB (multiple formats)
def parse_dob_safe(dob):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(dob), fmt).date()
        except (ValueError, TypeError):
            continue
    return None

# Parse Hire Date (multiple formats)
def parse_hire_date_safe(hire_date):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(hire_date), fmt).date()
        except (ValueError, TypeError):
            continue
    return None

def import_employees(request):
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('sowaf:employees')

    excel_file = request.FILES['excel_file']
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith('.csv'):
            decoded_file = excel_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader)  # Skip header row

            for row in reader:
                (
                first_name,last_name,gender,dob,nationality,nin_number,tin_number,profile_picture,phone_number,email_address,residential_address,emergency_person,emergency_contact,relationship,job_title,department,employment_type,status,hire_date,supervisor,salary,payment_frequency,payment_method,bank_name,bank_account,bank_branch,nssf_number,insurance_provider,taxable_allowances,intaxable_allowances,additional_notes
                ) = row
                
                dob = parse_dob_safe(dob)
                hire_date = parse_hire_date_safe(hire_date)
                profile_picture = profile_picture.strip() if profile_picture else ''


                employee = Newemployee.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    gender=gender,
                    dob=dob,
                    nationality=nationality,
                    nin_number=nin_number,
                    tin_number=tin_number,
                    profile_picture=profile_picture,
                    phone_number=str(phone_number).rstrip('.0') if phone_number else '',
                    email_address=email_address,
                    residential_address=residential_address,
                    emergency_person=emergency_person,
                    emergency_contact=emergency_contact,
                    relationship=relationship,
                    job_title=job_title,
                    department=department,
                    employment_type=employment_type,
                    status=status,
                    hire_date=hire_date,
                    supervisor=supervisor,
                    salary=salary,
                    payment_frequency=payment_frequency,
                    payment_method=payment_method,
                    bank_name=bank_name,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    nssf_number=nssf_number,
                    insurance_provider=insurance_provider,
                    taxable_allowances=taxable_allowances,
                    intaxable_allowances=intaxable_allowances,
                    additional_notes=additional_notes,
                )
                handle_profile_picture_upload(employee, profile_picture)

        elif file_name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                first_name,last_name,gender,dob,nationality,nin_number,tin_number,profile_picture,phone_number,email_address,residential_address,emergency_person,emergency_contact,relationship,job_title,department,employment_type,status,hire_date,supervisor,salary,payment_frequency,payment_method,bank_name,bank_account,bank_branch,nssf_number,insurance_provider,taxable_allowances,intaxable_allowances,additional_notes
                ) = row

                dob = parse_dob_safe(dob)
                hire_date = parse_hire_date_safe(hire_date)
                profile_picture = profile_picture.strip() if profile_picture else ''

                employee = Newemployee.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    gender=gender,
                    dob=dob,
                    nationality=nationality,
                    nin_number=nin_number,
                    tin_number=tin_number,
                    profile_picture=profile_picture,
                    phone_number=str(phone_number).rstrip('.0') if phone_number else '',
                    email_address=email_address,
                    residential_address=residential_address,
                    emergency_person=emergency_person,
                    emergency_contact=emergency_contact,
                    relationship=relationship,
                    job_title=job_title,
                    department=department,
                    employment_type=employment_type,
                    status=status,
                    hire_date=hire_date,
                    supervisor=supervisor,
                    salary=salary,
                    payment_frequency=payment_frequency,
                    payment_method=payment_method,
                    bank_name=bank_name,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    nssf_number=nssf_number,
                    insurance_provider=insurance_provider,
                    taxable_allowances=taxable_allowances,
                    intaxable_allowances=intaxable_allowances,
                    additional_notes=additional_notes,
                )
                handle_profile_picture_upload(employee, profile_picture)

        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect('sowaf:employees')

        messages.success(request, "employee data imported successfully.")
        return redirect('sowaf:employees')

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect('sowaf:employees')


# supplier view
def supplier(request):
    suppliers = Newsupplier.objects.all()
     
    return render(request, 'Supplier.html', {'suppliers':suppliers})

#add new supplier form view
def add_supplier(request):
    if request.method == 'POST':
        logo =request.FILES.get('logo')
        if logo:
            if not logo.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1048576:
                messages.error(request, "logo file size must not exceed 1MB.")
                return redirect(request.path)
        company_name = request.POST.get('company_name')
        supplier_type = request.POST.get('supplier_type')
        status = request.POST.get('status')
        contact_person = request.POST.get('contact_person')
        contact_position = request.POST.get('contact_position')
        contact = request.POST.get('contact')
        email = request.POST.get('email')
        open_balance = request.POST.get('open_balance')
        website = request.POST.get('website')
        address1 = request.POST.get('address1')
        address2 = request.POST.get('address2')
        city = request.POST.get('city')
        state = request.POST.get('state')
        zip_code = request.POST.get('zip_code')
        country = request.POST.get('country')
        bank = request.POST.get('bank')
        bank_account = request.POST.get('bank_account')
        bank_branch = request.POST.get('bank_branch')
        payment_terms = request.POST.get('payment_terms')
        currency = request.POST.get('currency')
        payment_method = request.POST.get('payment_method')
        tin = request.POST.get('tin')
        reg_number = request.POST.get('reg_number')
        tax_rate = request.POST.get('tax_rate')
        attachments =request.FILES.get('attachments')
        new_supplier = Newsupplier(logo=logo,company_name=company_name,supplier_type=supplier_type,status=status,contact_person=contact_person,contact_position=contact_position, contact=contact,email=email,open_balance=open_balance,website=website,address1=address1,address2=address2,city=city,state=state,zip_code=zip_code,country=country,bank=bank,bank_account=bank_account,bank_branch=bank_branch,payment_terms=payment_terms,currency=currency,payment_method=payment_method,tin=tin,reg_number=reg_number,tax_rate=tax_rate,attachments=attachments)

        # saving the data in the data base
        new_supplier.save()
         # adding save actions
        save_action = request.POST.get('save_action')
        if save_action == 'save&new':
            return redirect('add-suppliers')
        elif save_action == 'save&close':
            return redirect('sowaf:suppliers')
    return render(request, 'suppliers_entry_form.html', {})
# editing supplier information
def edit_supplier(request, pk):
    supplier = get_object_or_404(Newsupplier, pk=pk)

    if request.method == 'POST':
        supplier.company_name = request.POST.get('company_name', supplier.company_name)
        supplier.supplier_type = request.POST.get('supplier_type', supplier.supplier_type)
        supplier.status = request.POST.get('status', supplier.status)
        supplier.contact_person = request.POST.get('contact_person', supplier.contact_person)
        supplier.contact_position = request.POST.get('contact_position', supplier.contact_position)
        supplier.contact = request.POST.get('contact', supplier.contact)
        supplier.email = request.POST.get('email', supplier.email)
        supplier.open_balance = request.POST.get('open_balance', supplier.open_balance)
        supplier.website = request.POST.get('website', supplier.website)
        supplier.address1 = request.POST.get('address1', supplier.address1)
        supplier.address2 = request.POST.get('address2', supplier.address2)
        supplier.city = request.POST.get('city', supplier.city)
        supplier.state = request.POST.get('state', supplier.state)
        supplier.zip_code = request.POST.get('zip_code', supplier.zip_code)
        supplier.country = request.POST.get('country', supplier.country)
        supplier.bank = request.POST.get('bank', supplier.bank)
        supplier.bank_account = request.POST.get('bank_account', supplier.bank_account)
        supplier.bank_branch = request.POST.get('bank_branch', supplier.bank_branch)
        supplier.payment_terms = request.POST.get('payment_terms', supplier.payment_terms)
        supplier.currency = request.POST.get('currency', supplier.currency)
        supplier.payment_method = request.POST.get('payment_method', supplier.payment_method)
        supplier.tin = request.POST.get('tin', supplier.tin)
        supplier.bank_account = request.POST.get('bank_account', supplier.bank_account)
        supplier.bank_branch = request.POST.get('bank_branch', supplier.bank_branch)
        supplier.reg_number = request.POST.get('reg_number', supplier.reg_number)
        supplier.tax_rate = request.POST.get('tax_rate', supplier.tax_rate)

        # ✅ Only update logo if a new one is uploaded
        logo =request.FILES.get('logo')
        if logo:
            if not logo.name.lower().endswith('.png'):
                messages.error(request, "Only PNG files are allowed for the logo.")
                return redirect(request.path)
            if logo.size > 1048576:
                messages.error(request, "logo file size must not exceed 1MB.")
                return redirect(request.path)
            supplier.logo = logo
        if 'attachments' in request.FILES:
           supplier.attachments = request.FILES['attachments']

        supplier.save()
        return redirect('sowaf:suppliers')  # Or wherever your list view is

    return render(request, 'suppliers_entry_form.html', {'supplier': supplier})

    # Delete view
def delete_supplier(request, pk):
    supplier = get_object_or_404(Newsupplier, pk=pk)
    supplier.delete()
    return redirect('sowaf:suppliers')    

# importing suppliers
def download_suppliers_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers Template"

    headers = [
        'logo','company_name','supplier_type','status','contact_person','contact_position', 'contact','email','open_balance','website','address1','address2','city','state','zip_code','country','bank','bank_account','bank_branch','payment_terms','currency','payment_method','tin','reg_number','tax_rate',
    ]
    ws.append(headers)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="suppliers_template.xlsx"'
        return response
def handle_logo_upload(supplier, logo):
    if logo:
        image_path = os.path.join(settings.MEDIA_ROOT, 'uploads', logo)
        if os.path.exists(image_path):
            with open(image_path, 'rb') as f:
                supplier.logo.save(logo, File(f), save=True)
        else:
            messages.warning(None, f"Image file '{logo}' not found.")

def import_suppliers(request):
    if request.method != 'POST' or 'excel_file' not in request.FILES:
        messages.error(request, "No file uploaded.")
        return redirect('sowaf:suppliers')

    excel_file = request.FILES['excel_file']
    file_name = excel_file.name.lower()

    try:
        if file_name.endswith('.csv'):
            decoded_file = excel_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            next(reader)
            for row in reader:
                (
                    logo, company_name, supplier_type, status, contact_person, contact_position, contact, email, open_balance, website, address1, address2, city, state, zip_code, country, bank, bank_account,bank_branch,payment_terms,currency,payment_method,tin,reg_number,tax_rate,
                )
                
                logo = logo.strip() if logo else ''


                supplier = Newsupplier.objects.create(
                    logo=logo,
                    company_name=company_name,
                    supplier_type=supplier_type,
                    status=status,
                    contact_person=contact_person,
                    contact_position=contact_position, 
                    contact=contact,
                    email=email,
                    open_balance=open_balance,
                    website=website,
                    address1=address1,
                    address2=address2,
                    city=city,
                    state=state,
                    zip_code=zip_code,
                    country=country,
                    bank=bank,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    payment_terms=payment_terms,
                    currency=currency,
                    payment_method=payment_method,
                    tin=tin,
                    reg_number=reg_number,
                    tax_rate=tax_rate,
                )
                handle_logo_upload(supplier, logo)

        elif file_name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                (
                logo, company_name, supplier_type, status, contact_person, contact_position, contact, email, open_balance, website, address1, address2, city, state, zip_code, country, bank, bank_account,bank_branch,payment_terms,currency,payment_method,tin,reg_number,tax_rate,
                ) = row


                logo = logo.strip() if logo else ''

                supplier = Newsupplier.objects.create(
                    logo=logo,
                    company_name=company_name,
                    supplier_type=supplier_type,
                    status=status,
                    contact_person=contact_person,
                    contact_position=contact_position, 
                    contact=contact,
                    email=email,
                    open_balance=open_balance,
                    website=website,
                    address1=address1,
                    address2=address2,
                    city=city,
                    state=state,
                    zip_code=zip_code,
                    country=country,
                    bank=bank,
                    bank_account=bank_account,
                    bank_branch=bank_branch,
                    payment_terms=payment_terms,
                    currency=currency,
                    payment_method=payment_method,
                    tin=tin,
                    reg_number=reg_number,
                    tax_rate=tax_rate,
                )
                handle_logo_upload(supplier, logo)

        else:
            messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
            return redirect('sowaf:suppliers')

        messages.success(request, "supplier data imported successfully.")
        return redirect('sowaf:suppliers')

    except Exception as e:
        messages.error(request, f"Import failed: {str(e)}")
        return redirect('sowaf:suppliers')
# tasks view
def tasks(request):

    return render(request, 'tasks.html', {})
# taxes view
def taxes(request):

    return render(request, 'Taxes.html', {})


# expenses view
def expenses(request):

    return render(request, 'Expenses.html', {})

# millecious view
def miscellaneous(request):

    return render(request, 'Miscellaneous.html', {})
# reports view 
def reports(request):

    return render(request, 'Reports.html', {})